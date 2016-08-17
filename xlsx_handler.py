# coding: utf-8
import openpyxl
import xlrd
# import xlwt

from DataStructure.return_value import ReturnValue
from DataStructure.matrix import Matrix

class excel_lib_handler():
	def __init__(self):
		self._workbook = None
		self._current_sheet = None
		self._filename = ""


	def set_sheet_name(self, name):
		return ReturnValue(True, None)

	def set_cell(self, row, column, value):
		return ReturnValue(True, None)

	def save_workbook(self):
		return ReturnValue(True, None)

	def get_workbook(self):
		return ReturnValue(True, self._workbook)

	def new_workbook(self):
		return ReturnValue(True, None)

	def load_workbook(self):
		return ReturnValue(True, retval=self._workbook)


	def get_sheet_by_index(self, idx):
		return ReturnValue(True, None)


	def get_sheet_by_name(self, name):
		return ReturnValue(True, None)

	def get_sheet_names(self):
		"""
		:return [u'Sheet1', u'Sheet2', u'Sheet3']
		"""
		return ReturnValue(True, list())

	def get_col_number(self):
		"""
		:return: ReturnValue(bool, int)
		"""
		return ReturnValue(True, 0)

	def get_row_number(self):
		"""
		:return: ReturnValue(bool, int)
		"""
		return ReturnValue(True, 0)

	def get_row_values(self, row_idx):
		"""
		:return [u'1', u'56', u'O', u'', u'ANY', u'', u'', u'', u'SSLVPN', u'210.97.37.196', u'', u'SSLVPN', u'O', u'']
		"""
		retval = []
		return ReturnValue(True, retval)

	def get_cell_value(self, row_idx, cell_idx):
		return ReturnValue(True, None)



class openpyxl_handler(excel_lib_handler):
	def __init__(self, filename):
		excel_lib_handler.__init__(self)
		self._filename = filename

	def get_workbook(self):
		if self._workbook is None:
			return ReturnValue(False, None)
		return ReturnValue(True, self._workbook)

	def set_sheet_name(self, name):
		self._current_sheet.title = name

	def set_cell(self, row, column, value):
		self._current_sheet.cell(column=column, row=row, value=value)
		return ReturnValue(True, None)

	def new_workbook(self):
		self._workbook = openpyxl.Workbook(encoding="utf-8")
		self._workbook.create_sheet(title="Sheet1", index=0)
		self._current_sheet = self.get_sheet_by_index(0)['retval']
		return ReturnValue(True, None)

	def load_workbook(self):
		self._workbook = openpyxl.load_workbook(self._filename)
		self._current_sheet = self.get_sheet_by_index(0)['retval']
		return ReturnValue(True, retval=self._workbook)

	def save_workbook(self):
		self._workbook.save(self._filename)
		return ReturnValue(True, None)

	def get_sheet_by_index(self, idx):
		retval = self._workbook.worksheets[idx]
		return ReturnValue(True, retval)


	def get_sheet_by_name(self, name):
		retval = self._workbook.get_sheet_by_name(name)
		return ReturnValue(True, retval)


	def get_sheet_names(self):
		"""
		:return [u'Sheet1', u'Sheet2', u'Sheet3']
		"""
		retval = self._workbook.get_sheet_names()
		return ReturnValue(True, retval)

	def get_col_number(self):
		"""
		:return: ReturnValue(bool, int)
		"""
		retval = self._current_sheet.max_column
		return ReturnValue(True, retval)

	def get_row_number(self):
		"""
		:return: ReturnValue(bool, int)
		"""
		retval = self._current_sheet.max_row
		return ReturnValue(True, retval)

	def get_row_values(self, row_idx):
		"""
		:return [u'1', u'56', u'O', u'', u'ANY', u'', u'', u'', u'SSLVPN', u'210.97.37.196', u'', u'SSLVPN', u'O', u'']
		"""
		retval = []
		rows = self._current_sheet.rows
		for cell in rows[row_idx]:
			retval.append(cell.value)
		return ReturnValue(True, retval)

	def get_cell_value(self, row_idx, cell_idx):
		retval = self._current_sheet.rows[row_idx][cell_idx].value
		return ReturnValue(True, retval)


class xlrd_handler(excel_lib_handler):
	def __init__(self, filename):
		excel_lib_handler.__init__(self)
		self._filename = filename


	def get_workbook(self):
		if self._workbook is None:
			return ReturnValue(False, None)
		return ReturnValue(True, self._workbook)

	def load_workbook(self):
		self._workbook = xlrd.open_workbook(self._filename)
		self._current_sheet = self.get_sheet_by_index(0)['retval']
		return ReturnValue(True, retval=self._workbook)

	def get_sheet_names(self):
		"""
		:return [u'Sheet1', u'Sheet2', u'Sheet3']
		"""
		retval = self._workbook.sheet_names()
		return ReturnValue(True, retval)

	def get_sheet_by_index(self, idx):
		"""
		:return Object - <xlrd.sheet.Sheet>
		"""
		retval = self._workbook.sheet_by_index(idx)
		return ReturnValue(True, retval)

	def get_sheet_by_name(self, name):
		retval = self._workbook.sheet_by_name(name)
		return ReturnValue(True, retval)

	def get_col_number(self):
		"""
		:return: ReturnValue(bool, int)
		"""
		retval = self._current_sheet.ncols
		return ReturnValue(True, retval)

	def get_row_number(self):
		"""
		:return: ReturnValue(bool, int)
		"""
		retval = self._current_sheet.nrows
		return ReturnValue(True, retval)

	def get_row_values(self, row_idx):
		"""
		:return [u'1', u'56', u'O', u'', u'ANY', u'', u'', u'', u'SSLVPN', u'210.97.37.196', u'', u'SSLVPN', u'O', u'']
		"""
		retval = self._current_sheet.row_values(row_idx)
		retval = [v if v != "" else None for v in retval]
		return ReturnValue(True, retval)

	def get_cell_value(self, row_idx, col_idx):
		retval = self._current_sheet.cell_value(row_idx, col_idx)
		return ReturnValue(True, retval)






class ExcelHandler():
	def __init__(self, model, filename="unknown"):
		self.current_row = 0
		self.current_column = 0
		self._filename = filename
		self.model = model

		self._work_book = None

		self.read_handler = None

		self._filename = filename
		self.matrix = Matrix() # 메인 메트릭스

	def new_workbook(self, sheet_name="Sheet"):
		# 워크북 생성 xlsx 만 생성
		self._work_book = openpyxl_handler(filename=self._filename)
		self._work_book.new_workbook()


	def init_subtitle(self):
		pass

	def _insert_item(self, item):
		pass

	def convert(self):
		"""
		model의 아이템을 순서대로 읽고 기록
		"""
		# 최상단 필드 이름 기록
		self.init_subtitle()

		# cell 시작 위치 초기화
		self.current_column = 1
		self.current_row = 2

		# model에서 아이템을 하나씩 꺼내어 기록
		for item in self.model:
			rtv = self._insert_item(item)
			if not rtv["state"]:
				raise Exception

			m = rtv["retval"]['matrix']
			self.matrix.set_matrix(self.current_row - 1, self.current_column - 1, m)
			self.current_row += rtv["retval"]['row']
			self.current_row += 1

		self.write_matrix(self.matrix)
		self.write_screen(self.matrix)

	def load(self):
		file_ext = self._filename[self._filename.rfind('.'):]
		if file_ext == ".xls":
			self._work_book = xlrd_handler(self._filename)
			self._work_book.load_workbook()
		elif file_ext == ".xlsx":
			self._work_book = openpyxl_handler(self._filename)
			self._work_book.load_workbook()
		else:
			raise TypeError

	def read(self, row, col):
		return self._work_book.get_cell_value(row, col)['retval']


	def set_cell_value(self, column, row, value):
		""""(row, col)에 데이터 삽입"""
		self._work_book.set_cell(column=column, row=row, value=value)

	def merge(self, s_row, e_row, s_col, e_col):
		pass

	def write_matrix(self, matrix):
		for row, row_data in enumerate(matrix, 0):
			for col, col_data in enumerate(row_data, 0):
				if matrix[row][col] is None: continue
				self.set_cell_value(col + 1, row + 1, matrix[row][col])
		return ReturnValue(True, None)

	def write_screen(self, matrix):
		temp = []
		for row, row_data in enumerate(matrix, 0):
			for col, col_data in enumerate(row_data, 0):
				d = matrix[row][col]
				temp.append("{:d}: {:^10s} | ".format(col, str(d if d is not None else '...'*3)[:10]))
			temp.append("\n")
		print ''.join(temp)
		return ReturnValue(True, None)


	def set_sheet_name(self, name):
		self._work_book.set_sheet_name(name)

	def get_sheet_by_name(self, name):
		return self._work_book.sheet_by_name(name)['retval']

	def get_row_values(self, row_idx):
		return self._work_book.get_row_values(row_idx)

	@property
	def sheets_names(self):
		return self._work_book.get_sheet_names()['retval']

	@property
	def end_col(self):
		return self._work_book.get_col_number()['retval']

	@property
	def end_row(self):
		return self._work_book.get_row_number()['retval']

	def save(self):
		state = self._work_book.save_workbook()['state']
		return ReturnValue(state, None)

	@staticmethod
	def get_merged_row_cood(data):
		"""
		엑셀에서 머지된 row 를 구분 하여 좌표를 돌려준다.
		:param data: [[]]
		:return: ((0, 5), (5, 6))

		>>> data = [
		>>>	["A",0,1,2,3,4],
		>>> [None,0,1,2,3,4],
		>>> [None,0,1,2,3,4],
		>>> ["B",1,1,2,3,4],
		>>> [None,1,1,2,3,4],
		>>> [None,1,1,2,3,4],
		>>> ["C",1,1,2,3,4],
		>>> [None,1,1,2,3,4],
		>>> [None,1,1,2,3,4],
		>>> [None,1,1,2,3,4],
		>>> [None,1,1,2,3,4],
		>>> ]
		>>> print data
		"""
		if not data[0]:
			return ReturnValue(False, None)


		coodination = []
		start = 0
		end = 0
		for i, d in enumerate(data):
			end = i
			if d[0] and i != 0:
				coodination.append((start, end))
				start = i

		coodination.append((start, end + 1))

		return ReturnValue(True, tuple(coodination))


if __name__ == "__main__":
	ehd = ExcelHandler(None, "test.xls")
	ehd.load()

	data_map = []
	for i in range(4, ehd.end_row):
		data_map.append(ehd.get_row_values(i)['retval'])

	coodination = ehd.get_merged_row_cood(data_map)['retval']

	items = []
	for c in coodination:
		src = []
		dst = []
		svc = []
		for d in data_map[c[0]:c[1]]:
			src.append(d[3:7])
			dst.append(d[7:10])
			svc.append(d[11])
		items.append({"src":src,"dst":dst,"svc":svc})

	for item in items:
		print item






